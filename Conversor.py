import pdfplumber
import os
import re
import sys
from tkinter import Tk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

# =============================================================================
# CONFIGURAÇÕES DO MODELO (ALTERE AQUI PARA CADA NOVO CLIENTE/PDF)
# =============================================================================

# 1. Tabela de Conversão (De: Código do Fornecedor -> Para: Seu Código Interno)
TABELA_DE_PARA = {
    "COD-FORN-001": "MEU-ITEM-AAA",
    "COD-FORN-002": "MEU-ITEM-BBB",
    # Adicione quantos códigos precisar aqui
}

# 2. Palavras-chave para identificar a linha do "Nome do Produto/Material" no PDF
PALAVRAS_CHAVE_PRODUTO = ["MATERIAL", "PEÇA", "ITEM", "DESCRIÇÃO", "PART NUMBER"]

# 3. Expressão Regular para capturar o código do produto (Padrão: 0-000-000-000)
# Se o padrão mudar, altere esta Regex.
REGEX_CODIGO_PRODUTO = r'(\d[\d\s-]{7,15}\d[a-zA-Z]?)'

# 4. Expressão Regular para capturar Datas (Padrão: DD.MM.AAAA)
REGEX_DATA = r'(\d{2}\.\d{2}\.\d{4})'

# 5. Nomes de Saída
NOME_DO_RELATORIO = "RELATORIO_PROCESSAMENTO.xlsx"
NOME_DA_ABA = "Dados_Extraidos"
# =============================================================================

def processar_pdfs():
    root = Tk()
    root.withdraw()
    root.wm_attributes("-topmost", 1)
    
    arquivos = filedialog.askopenfilenames(title="SELECIONE OS PDFS PARA CONVERSÃO", filetypes=[("PDF", "*.pdf")])
    if not arquivos: return

    # Define pasta de saída (mesma pasta do executável ou script)
    diretorio_atual = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    pasta_saida = os.path.join(diretorio_atual, "Relatorios_Gerados")
    if not os.path.exists(pasta_saida): os.makedirs(pasta_saida)
    
    caminho_excel = os.path.join(pasta_saida, NOME_DO_RELATORIO)

    dados_acumulados = {}
    datas_colunas = set()

    # --- 1. CARREGAR HISTÓRICO EXISTENTE (Para não perder dados anteriores) ---
    if os.path.exists(caminho_excel):
        try:
            wb_ref = load_workbook(caminho_excel, data_only=True)
            ws_ref = wb_ref.active
            headers = [str(cell.value) for cell in ws_ref[1]]
            # Identifica colunas que são datas
            idx_datas = {h: i for i, h in enumerate(headers) if re.match(r'\d{2}\.\d{2}\.\d{4}', h)}
            
            for row in ws_ref.iter_rows(min_row=2, values_only=True):
                if not row[0] or "TOTAL" in str(row[0]).upper(): continue
                item_id = str(row[0])
                if item_id not in dados_acumulados: dados_acumulados[item_id] = {"valores": {}}
                for h, idx in idx_datas.items():
                    if row[idx] is not None:
                        dados_acumulados[item_id]["valores"][h] = float(row[idx])
                        datas_colunas.add(h)
            wb_ref.close()
        except: pass

    # --- 2. LER E EXTRAIR DADOS DOS NOVOS PDFs ---
    for caminho in arquivos:
        try:
            with pdfplumber.open(caminho) as pdf:
                id_documento = "Desconhecido"
                for pagina in pdf.pages:
                    texto = pagina.extract_text()
                    if not texto: continue
                    linhas = texto.split('\n')
                    
                    # Busca o Identificador do Produto (Baseado nas Palavras-Chave)
                    for linha in linhas:
                        if any(p in linha.upper() for p in PALAVRAS_CHAVE_PRODUTO):
                            match = re.search(REGEX_CODIGO_PRODUTO, linha)
                            if match:
                                # Limpa o código (remove espaços e padroniza hífens)
                                id_documento = match.group(1).strip().replace(" ", "-")
                                break
                    
                    # Busca Datas e Valores na mesma linha
                    for linha in linhas:
                        match_data = re.search(REGEX_DATA, linha)
                        if match_data:
                            data_v = match_data.group(1)
                            partes = linha.split()
                            
                            # Lógica Genérica: O valor costuma ser o último item da linha
                            raw_val = partes[-1]
                            # Limpa formatação de moeda (Ex: 1.250,50 -> 1250.50)
                            limpo = raw_val.replace('.', '').replace(',', '.')
                            limpo = re.sub(r'[^\d.]', '', limpo)
                            
                            try:
                                valor_num = float(limpo)
                                if id_documento not in dados_acumulados: 
                                    dados_acumulados[id_documento] = {"valores": {}}
                                
                                # Salva o valor na data correspondente
                                dados_acumulados[id_documento]["valores"][data_v] = valor_num
                                datas_colunas.add(data_v)
                            except: continue
        except: continue

    # --- 3. GERAR O EXCEL FINAL ---
    wb = Workbook()
    ws = wb.active
    ws.title = NOME_DA_ABA
    
    # Ordena as colunas por data
    datas_ord = sorted(list(datas_colunas), key=lambda x: datetime.strptime(x, '%d.%m.%Y'))
    ws.append(["ID Original", "Código Interno"] + datas_ord)

    for item_id, info in dados_acumulados.items():
        # Aplica a Tabela de De-Para definida no topo
        linha = [item_id, TABELA_DE_PARA.get(item_id, "NÃO MAPEADO")]
        for d in datas_ord:
            linha.append(info["valores"].get(d, 0))
        ws.append(linha)

    # Adiciona linha de Totais
    last_row = ws.max_row + 1
    ws.append(["TOTAL GERAL", ""])
    for col_idx in range(3, len(datas_ord) + 3):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.cell(row=last_row, column=col_idx).value = f"=SUM({col_letter}2:{col_letter}{last_row-1})"

    # --- 4. ESTILIZAÇÃO VISUAL ---
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if cell.row == 1: 
                cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="D3D3D3")
            if "TOTAL GERAL" in str(row[0].value):
                cell.font = Font(bold=True, color="FF0000")

    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 18

    try:
        wb.save(caminho_excel)
        messagebox.showinfo("CONCLUÍDO", f"O arquivo {NOME_DO_RELATORIO} foi gerado com sucesso!")
    except:
        messagebox.showerror("ERRO", "Não foi possível salvar. O arquivo Excel está aberto?")

if __name__ == "__main__":
    processar_pdfs()